from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout
from payback.models import Technoplayer
from .models import *
import xlrd
import os
from django.conf import settings
from openpyxl import load_workbook
from io import BytesIO
from payback.forms import *
# from django.conf.settings import PROJECT_ROOT

# Create your views here.
def opening(request):
    return render(request, 'opening.html')


def firstyear(request):
    return render(request, 'firstyear.html')


def kenken(request):
    return render(request, 'kenken.html')


def slidepuzzle(request):
    return render(request, 'slidingpuzzle.html')


def slider(request):
    return render(request, 'slider.html')


def tangram(request):
    return render(request, 'tangram.html')


def secondyear(request):
    technoplayer = Technoplayer.objects.filter(user=request.user).first()
    loan = technoplayer.loan
    connection = technoplayer.connection
    happiness= technoplayer.happiness
    focus= technoplayer.focus
    return render(request, 'secondyear.html',{'loan':loan,'connection':connection,'happiness':happiness})


def thirdyear(request):
    return render(request, 'thirdyear.html')


def fourthyear(request):
    return render(request, 'fourthyear.html')


def graduation(request):
    return render(request, 'graduation.html')


def game(request):
    return render(request, 'game.html')


def mastermind(request):
    return render(request, 'mastermind.html')


def crossword(request):
    return render(request, 'crossword_begining.html')


def mysteryroom(request):
    return render(request, 'mystery_room.html')


def firstyear_submission(request):
    technoplayer = Technoplayer.objects.filter(user=request.user).first()
    if request.method == "POST":
        focus = request.POST.get('focus')
        happiness = request.POST.get('happiness')
        connection = request.POST.get('connection')
        loan = request.POST.get('loan')
        print(focus)
        if technoplayer is not None:
            technoplayer.happiness = happiness
            technoplayer.connection = connection
            technoplayer.focus = focus
            technoplayer.loan = loan
            technoplayer.save()
        else:
            technoplayer = Technoplayer()
            technoplayer.user = request.user
            technoplayer.happiness = happiness
            technoplayer.connection = connection
            technoplayer.focus = focus
            technoplayer.loan = loan
            technoplayer.save()

        # thirdyear = Thirdyear(puzzle_score=puzzle_score, age_sum=agesum_solved, letter_sum=letter_sum_solved)
        # thirdyear.save();
        data = "Save Successfully"
        return JsonResponse(data, safe=False)

    return HttpResponse("get method")

# def login(request):
#     if request.method == 'POST':
#         rollno = (request.POST['rollno'])
#         password = (request.POST['password'])
#         technoplayer = Technoplayer.objects.get(rollno=rollno)
#         # user = authenticate(rollno = rollno, password=password)
#         if technoplayer is not None:
#             print('user logged in')
#             return JsonResponse("Login Successful",safe=False)

def techo_login(request):
    if request.user.is_authenticated:
        return redirect(request.GET.get('next','/firstyear'))

    if request.method=="POST":
        technouser = User.objects.filter(roll_no=request.POST['roll_no']).first()
        if technouser is None:
            return render(request, 'login.html', {"messages":[["text-danger","Roll Number Not Found."]]})
        username = technouser.username;
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            print("Used logged in!")
            return redirect(request.GET.get('next','/firstyear'))
        else:
            return render(request, 'login.html',{"messages":[["text-danger","Invalid Credentials."]]})
    return render(request, 'login.html',)



def logout_view(request):
    logout(request)
    return redirect('/')

def secondyear_submission(request):
    if request.method == "POST":
        focus = request.POST.get('focus')
        print(focus)
        happiness = request.POST.get('happiness')
        connection = request.POST.get('connection')
        loan = request.POST.get('loan')
        technoplayer = Technoplayer.objects.filter(user=request.user)
        if technoplayer != None:
            technoplayer.happiness = happiness
            technoplayer.connection = connection
            technoplayer.focus = focus
            technoplayer.loan = loan
            technoplayer.save()

        # thirdyear = Thirdyear(puzzle_score=puzzle_score, age_sum=agesum_solved, letter_sum=letter_sum_solved)
        # thirdyear.save();
        data = "Save Successfully"
        return JsonResponse(data, safe=False)

    return HttpResponse("get method")


def thirdyear_submission(request):
    if request.method == "POST":
        focus = request.POST.get('focus')
        print(focus)
        happiness = request.POST.get('happiness')
        connection = request.POST.get('connection')
        loan = request.POST.get('loan')
        technoplayer = Technoplayer.objects.filter(user=request.user)
        if technoplayer != None:
            technoplayer.happiness = happiness
            technoplayer.connection = connection
            technoplayer.focus = focus
            technoplayer.loan = loan
            technoplayer.save()

        # thirdyear = Thirdyear(puzzle_score=puzzle_score, age_sum=agesum_solved, letter_sum=letter_sum_solved)
        # thirdyear.save();
        data = "Save Successfully"
        return JsonResponse(data, safe=False)

    return HttpResponse("get method")


def fourthyear_submission(request):
    if request.method == "POST":
        focus = request.POST.get('focus')
        print(focus)
        happiness = request.POST.get('happiness')
        connection = request.POST.get('connection')
        loan = request.POST.get('loan')
        technoplayer = Technoplayer.objects.filter(user=request.user)
        if technoplayer != None:
            technoplayer.happiness = happiness
            technoplayer.connection = connection
            technoplayer.focus = focus
            technoplayer.loan = loan
            technoplayer.save()

        # thirdyear = Thirdyear(puzzle_score=puzzle_score, age_sum=agesum_solved, letter_sum=letter_sum_solved)
        # thirdyear.save();
        data = "Save Successfully"
        return JsonResponse(data, safe=False)

    return HttpResponse("get method")

def append_user(excel_data):

    wb = load_workbook(filename=BytesIO(excel_data))
    sheet = wb.active
    # a1 = sheet['A1']
    r = 10
    col = 5
    for i in  range(r):
        name_val = sheet.cell(row=i+2, column=1).value
        print(name_val)
        name = name_val.split()
        first_name = name[0]
        last_name = name[1]
        username = str(first_name.lower()+last_name.lower())
        contact = sheet.cell(row=i+2, column=2).value
        email = sheet.cell(row=i+2, column=3).value
        roll = str(sheet.cell(row=i+2, column=4).value)
        password = str(sheet.cell(row=i+2, column=5).value)
        user = User.objects.create_user(username=username, email=email, password=password)
        user.first_name = first_name
        user.last_name = last_name
        user.roll_no = roll
        user.contact = contact
        user.save()

    # for i in range(sheet.nrows):
    #     print(sheet.cell_value(i,0))

def get_sheet(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file_in_memory = request.FILES['file'].read()
            # wb = load_workbook(filename=BytesIO(file_in_memory))
            append_user(file_in_memory)

            return HttpResponse("Read Successfully")


    else:
        form = UploadFileForm()
        return render(request, 'upload.html', {'form': form})